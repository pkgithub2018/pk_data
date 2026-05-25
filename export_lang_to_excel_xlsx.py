#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Export lang_la.php translation file to Excel format (.xlsx)
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Read the PHP file
input_file = r'c:\xampp\htdocs\php-bin\lang_la.php'
output_excel = r'c:\xampp\htdocs\lang_la_translations.xlsx'

translations = []

print("Reading PHP file...")
with open(input_file, 'r', encoding='utf-8') as f:
    content = f.read()
    
    # Find all translation pairs using regex
    # Pattern matches: 'key' => 'value',
    pattern = r"'([^']+)'\s*=>\s*'([^']*(?:''[^']*)*)'"
    
    matches = re.findall(pattern, content)
    
    for key, value in matches:
        translations.append([key, value])

print(f"Found {len(translations)} translations")

# Create Excel workbook
print(f"Creating Excel file: {output_excel}")
wb = Workbook()
ws = wb.active
ws.title = "Lao Translations"

# Style the header row
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# Write header
ws['A1'] = 'English Key'
ws['B1'] = 'Lao Translation'

# Apply header styles
for cell in ['A1', 'B1']:
    ws[cell].font = header_font
    ws[cell].fill = header_fill
    ws[cell].alignment = header_alignment

# Write data
for idx, (key, value) in enumerate(translations, start=2):
    ws[f'A{idx}'] = key
    ws[f'B{idx}'] = value

# Adjust column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 50

# Freeze the header row
ws.freeze_panes = 'A2'

# Save the workbook
wb.save(output_excel)

print(f"✓ Successfully exported {len(translations)} translations to Excel!")
print(f"File location: {output_excel}")
print(f"\nYou can now open this file in Microsoft Excel or LibreOffice Calc.")
